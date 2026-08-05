#!/usr/bin/env python3
"""Read-only integrity and completeness checks for the recovered Fiji lineage."""

from __future__ import annotations

import csv
import hashlib
import math
import sys
import zipfile
from pathlib import Path


PACKAGE = Path(__file__).resolve().parents[1]
REPO = PACKAGE.parent
EVIDENCE = PACKAGE / "data_sources/evidence"

EXPECTED_HASHES = {
    "v25_lineage/archives/Fiji_CLEWs_Global_source_raw_pre_tracking_2026-07-25.zip": "6f53a3788a374ad86ebbc7ae2a41df5a8bc47d002aacc6a64d3b3b88d112c51d",
    "v25_lineage/archives/Fiji_v2_v2.5.0_HANDOFF_2026-07-29.zip": "3dfe81379dc249cde99886be52923787064a7caefc4bde915f6b1eb4546ca729",
    "v25_lineage/ARCHIVE_MANIFEST.csv": "6ec72104599753f3f48f3e63b8aa95972cada50259804ce75e967e3f47bcca23",
    "raw_clews_lineage/gaez_input_tables/GAEZ_cwd_High_Input.csv": "900596e1d956adc18e2c77cbc3a07e50c7363f72897f350b88c6a2a2e7db3e9a",
    "raw_clews_lineage/gaez_input_tables/GAEZ_cwd_Low_Input.csv": "6817b8d2561fac0bb7bfc38ba8076a50fd4d9e449815b78dfed8a9bb827afcd1",
    "raw_clews_lineage/gaez_input_tables/GAEZ_evt_High_Input.csv": "5e1df43709dfc16d94e690c0bacd76e8829d8e86cf747322a52467c7296e7cbc",
    "raw_clews_lineage/gaez_input_tables/GAEZ_evt_Low_Input.csv": "c931c0d6c6dffff7936f3835f872560576ca6f739c2c99cea713bac92220a874",
    "raw_clews_lineage/gaez_input_tables/GAEZ_yld_High_Input.csv": "47121388a94a61cd7158e9f2ad9474d3480d93b2aa292e154868760bcf84a531",
    "raw_clews_lineage/gaez_input_tables/GAEZ_yld_Low_Input.csv": "051c31b303930775bbe12229fd59b9257eed5d36b5a01c3249d4655e498a8beb",
    "raw_clews_lineage/gaez_input_tables/GAEZ_FJI_RASTER_CACHE_MANIFEST.csv": "61e7d82ddee3780747031411228f5acfca74a263f376f20d5c26749dee04a113",
    "raw_clews_lineage/gaez_base_rasters/GAEZ_FJI_BUNDLED_BASE_RASTERS.csv": "368266180b27fbb5ad9a7744817310f68ed9f339710bbef1f5bbef31e9390658",
    "raw_clews_lineage/gaez_base_rasters/LCType_ncb.tif": "17b77aa5fd7b56a4570119026034c0f3137b9e906c741c684556f762c34a30db",
    "raw_clews_lineage/gaez_base_rasters/precipitation prc.tif": "c72b8cec38ea31755b9d6f02e518571f41f7d39ae2f5913ee64a14077079d83b",
    "raw_clews_lineage/faostat/Crop_code.csv": "2555b2fd4a87a9612db88f526d050a6449ed0fd58fa171ac8503c59bd42d2f61",
    "raw_clews_lineage/faostat/FAOSTAT_2020.csv": "082da2e1c708ea6f9837b021143eb193d9f240b0700be367d80ab91cfeb259b9",
    "raw_clews_lineage/faostat/FAOSTAT_production_2020.csv": "51fe2f975b0d57ddcca0e4da860b12c89b7aa9a9daa41b3fe13df77fd746a14f",
    "raw_clews_lineage/faostat/FAOSTAT_FJI_2020_SELECTION.csv": "170de102e7c150e7230775e535be7a8bce7942c6a36d423734c55e0210e77582",
    "raw_clews_lineage/ssp2/iamc_db_POP_Countries.xlsx": "e013c76eacd95a2e0f6275446e9d1e71bdfb9556eab1c6619a5b0d7914ecab35",
    "raw_clews_lineage/ssp2/SSP2_FJI_SELECTED_ROW.csv": "77ca5cd1778dc546a1af6b90ca03951a59de8c90d6afedb5cc7957996bd86afb",
    "raw_clews_lineage/ssp2/SSP2_FJI_ANNUAL_INDEX_2021_2050.csv": "57c54345a1978ef5e7eaa2b9d0545e85e24c070fe0b1c480dfc433949ab5347a",
}

LIVE_HASHES = {
    "genData.json": "8987e3ae95a1a9b594278282a3930a822c56ad82ebeb3008ea85e2134baba8c4",
    "RYC.json": "f75f5ef8b873bd59acc4031ee9d87378947686f1660684230c73dc0f2c5709a4",
    # Post-v2.9 Fisheries activity-ceiling source hash from
    # validation/fisheries_bounds_v29_final.json.
    "RYT.json": "ff8e8c9598c062767cc0bd76c4b0c1c7ff9d72eafff587f7d630d14ade5957cd",
    "RT.json": "c3fcb7aded5d2c0edaad454e6a01abe445d43b5157438802fb23081d4085c075",
    "RYTCM.json": "912ce0de12b2c319587dde5d85784fcd53bc2e5ae92ab083271493075ff2a662",
    "RYTM.json": "58a1a3fbff88e95d8f3cefcc1965628f254580db80ec0da27fb8aa335b8bfe59",
    "RYTTs.json": "ad7113924ee75accc23d93a3a6719b618e8a6d5b4923190733bdc402eb1c2cd4",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as stream:
        return list(csv.DictReader(stream))


def check(condition: bool, message: str, failures: list[str]) -> None:
    if not condition:
        failures.append(message)


def main() -> int:
    failures: list[str] = []

    for relative, expected in EXPECTED_HASHES.items():
        path = EVIDENCE / relative
        check(path.is_file(), f"missing evidence file: {relative}", failures)
        if path.is_file():
            check(sha256(path) == expected, f"hash mismatch: {relative}", failures)

    raw_zip = EVIDENCE / "v25_lineage/archives/Fiji_CLEWs_Global_source_raw_pre_tracking_2026-07-25.zip"
    v25_zip = EVIDENCE / "v25_lineage/archives/Fiji_v2_v2.5.0_HANDOFF_2026-07-29.zip"
    with zipfile.ZipFile(raw_zip) as archive:
        check(len(archive.infolist()) == 190, "raw archive should contain 190 entries", failures)
        names = set(archive.namelist())
        for suffix in (
            "config/config.yaml",
            "UPSTREAM.lock",
            "overrides/workflow/scripts/clewsy.py",
            "overrides/workflow/submodules/CLEWs_GAEZ/workflow/scripts/collect.py",
        ):
            check(any(name.endswith(suffix) for name in names), f"raw archive lacks {suffix}", failures)

    with zipfile.ZipFile(v25_zip) as archive:
        check(len(archive.infolist()) == 321, "v2.5 handoff should contain 321 entries", failures)
        names = set(archive.namelist())
        for suffix in (
            "data_sources/ASSUMPTIONS.csv",
            "data_sources/CALCULATIONS.csv",
            "data_sources/MODEL_DATA_MAP.csv",
            "scripts/build_fiji_v2.py",
            "scripts/apply_fiji_phase1b_public_water.py",
            "scripts/apply_fiji_phase1c_bottom_up_demand.py",
            "scripts/apply_fiji_phase1d_cane_bagasse.py",
            "data_sources/evidence/external/EFL_2024_Annual_Report.pdf",
            "data_sources/evidence/external/Fiji_REI_Investment_Plan_2023.pdf",
            "data_sources/evidence/petroleum/UNSD_Energy_DSD_Matrix.xlsm",
        ):
            check(any(name.endswith(suffix) for name in names), f"v2.5 handoff lacks {suffix}", failures)

    gaez = rows(EVIDENCE / "raw_clews_lineage/gaez_input_tables/GAEZ_FJI_RASTER_CACHE_MANIFEST.csv")
    check(len(gaez) == 72, "GAEZ manifest should contain 72 crop-raster rows", failures)
    check({row["model_crop_code"] for row in gaez} == {"CAS", "CON", "PTS", "RCP", "SGC", "YAM"}, "GAEZ crop codes differ", failures)
    check(sum(row["input_level"] == "High" for row in gaez) == 36, "GAEZ high-input count differs", failures)
    check(sum(row["input_level"] == "Low" for row in gaez) == 36, "GAEZ low-input count differs", failures)
    check(len({row["cache_filename"] for row in gaez}) == 72, "GAEZ cache filenames are not unique", failures)
    check(all(row["download_url"].startswith("https://") for row in gaez), "GAEZ URL missing or non-HTTPS", failures)

    fao = rows(EVIDENCE / "raw_clews_lineage/faostat/FAOSTAT_FJI_2020_SELECTION.csv")
    check(len(fao) == 10, "FAOSTAT Fiji selection should contain 10 rows", failures)
    check(fao[0]["Item"] == "Sugar cane" and fao[0]["Value_harvested_area"] == "38000", "FAOSTAT lead selection differs", failures)
    check(all(row["_source_csv_line_harvested_area"] and row["_source_csv_line_production"] for row in fao), "FAOSTAT source-line locator missing", failures)

    ssp_row = rows(EVIDENCE / "raw_clews_lineage/ssp2/SSP2_FJI_SELECTED_ROW.csv")
    check(len(ssp_row) == 1, "SSP2 selected-row extract should contain one row", failures)
    selected = ssp_row[0]
    check(
        selected.get("sheet") == "data"
        and selected.get("excel_row") == "553"
        and selected.get("Model") == "IIASA-WiC POP"
        and selected.get("Scenario") == "SSP2"
        and selected.get("Region") == "FJI"
        and selected.get("Variable") == "Population",
        "SSP2 row identity differs",
        failures,
    )
    annual = rows(EVIDENCE / "raw_clews_lineage/ssp2/SSP2_FJI_ANNUAL_INDEX_2021_2050.csv")
    check([int(row["year"]) for row in annual] == list(range(2021, 2051)), "SSP2 annual series does not cover 2021-2050", failures)
    check(math.isclose(float(annual[0]["population_index_2021"]), 1.0), "SSP2 annual index is not normalized at 2021", failures)

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(EVIDENCE / "raw_clews_lineage/ssp2/iamc_db_POP_Countries.xlsx", read_only=True, data_only=True)
        values = list(next(workbook["data"].iter_rows(min_row=553, max_row=553, values_only=True)))
        workbook.close()
        check(values[:5] == ["IIASA-WiC POP", "SSP2", "FJI", "Population", "million"], "SSP2 workbook row identity differs", failures)
        years = [str(year) for year in range(2010, 2101, 5)]
        for offset, year in enumerate(years, start=6):
            check(math.isclose(float(selected[year]), float(values[offset - 1]), rel_tol=0, abs_tol=1e-12), f"SSP2 workbook value differs for {year}", failures)
    except ImportError:
        failures.append("openpyxl is required to verify the SSP2 workbook row")

    live_candidates = (
        REPO / "WebAPP/DataStorage/Fiji_v2.9",
        REPO.parent / "MUIOGO/WebAPP/DataStorage/Fiji_v2.9",
    )
    live = next((candidate for candidate in live_candidates if candidate.is_dir()), None)
    if live is not None:
        for name, expected in LIVE_HASHES.items():
            path = live / name
            check(path.is_file(), f"missing live Fiji_v2.9 input: {name}", failures)
            if path.is_file():
                check(sha256(path) == expected, f"live Fiji_v2.9 input changed: {name}", failures)

    if failures:
        print(f"Fiji v2.5 lineage recovery: FAIL ({len(failures)} failure(s))")
        for failure in failures:
            print(f"  - {failure}")
        return 1

    print("Fiji v2.5 lineage recovery: PASS")
    print(f"  evidence hashes: {len(EXPECTED_HASHES)}")
    print("  archives: 190-entry raw build; 321-entry v2.5 handoff")
    print("  GAEZ: 72 crop-raster requests + 2 retained base rasters")
    print("  FAOSTAT: 10 exact Fiji 2020 selections")
    print("  SSP2: data!A553:Y553 and annual 2021-2050 index")
    if live is None:
        print("  live Fiji_v2.9 inputs: not present; evidence-only verification completed")
    else:
        print(f"  live Fiji_v2.9 inputs unchanged: {len(LIVE_HASHES)} authoritative JSON files")
    return 0


if __name__ == "__main__":
    sys.exit(main())
