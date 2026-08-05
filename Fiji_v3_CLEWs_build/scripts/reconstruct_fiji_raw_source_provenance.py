#!/usr/bin/env python3
"""Reconstruct Fiji raw-build GAEZ, FAOSTAT and SSP2 source selections."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
EVIDENCE = ROOT / "data_sources" / "evidence" / "raw_clews_lineage"
GAEZ = EVIDENCE / "gaez_input_tables"
FAO = EVIDENCE / "faostat"
SSP = EVIDENCE / "ssp2"


def write_rows(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def reconstruct_faostat() -> tuple[list[str], list[str]]:
    harvested = pd.read_csv(FAO / "FAOSTAT_2020.csv", dtype={"Item Code (CPC)": str})
    production = pd.read_csv(
        FAO / "FAOSTAT_production_2020.csv", dtype={"Item Code (CPC)": str}
    )
    codes = pd.read_csv(FAO / "Crop_code.csv", dtype=str)
    code_by_item = dict(zip(codes["Name"], codes["Code"]))

    fiji = harvested[harvested["Area"] == "Fiji"].copy()
    fiji["_physical_line"] = fiji.index + 2
    top = fiji.nlargest(10, "Value").copy()
    items = top["Item"].tolist()
    main_items, other_items = items[:5], items[5:]
    main_codes = list(dict.fromkeys(code_by_item.get(item) for item in main_items))
    other_codes = [
        code
        for code in dict.fromkeys(code_by_item.get(item) for item in other_items)
        if code and code not in set(main_codes)
    ]

    prod = production[production["Area"] == "Fiji"].copy()
    prod["_physical_line"] = prod.index + 2
    prod = prod[prod["Item"].isin(items)].set_index("Item")
    modelled = {"CAS", "CON", "OTH", "SGC", "YAM"}
    output: list[dict[str, object]] = []
    for _, area in top.iterrows():
        item = area["Item"]
        matching = prod.loc[item]
        if isinstance(matching, pd.DataFrame):
            matching = matching.iloc[0]
        proxy = code_by_item.get(item, "")
        group = "OTH" if item in other_items or proxy not in modelled else proxy
        output.append(
            {
                "Item Code (CPC)": area["Item Code (CPC)"],
                "Item": item,
                "gaez_proxy_code": proxy,
                "model_output_group": group,
                "Value_harvested_area": area["Value"],
                "Unit_harvested_area": area["Unit"],
                "Flag_harvested_area": area["Flag"],
                "Flag Description_harvested_area": area["Flag Description"],
                "_source_csv_line_harvested_area": int(area["_physical_line"]),
                "Value_production": matching["Value"],
                "Unit_production": matching["Unit"],
                "Flag_production": matching["Flag"],
                "Flag Description_production": matching["Flag Description"],
                "_source_csv_line_production": int(matching["_physical_line"]),
            }
        )
    write_rows(FAO / "FAOSTAT_FJI_2020_SELECTION.csv", list(output[0]), output)
    return main_codes, other_codes


def reconstruct_gaez(main_codes: list[str], other_codes: list[str]) -> None:
    codes = pd.read_csv(FAO / "Crop_code.csv", dtype=str)
    code_by_gaez = dict(zip(codes["GAEZ_name"], codes["Code"]))
    final: dict[str, dict[str, object]] = {}
    for metric in ("yld", "cwd", "evt"):
        for level in ("High", "Low"):
            table = f"GAEZ_{metric}_{level}_Input.csv"
            frame = pd.read_csv(GAEZ / table, dtype=str)
            frame["_physical_line"] = frame.index + 2
            frame["New Crop"] = frame["Crop"].map(code_by_gaez).fillna("Nan")
            frame["New Water Supply"] = frame["Water Supply"].map(
                lambda value: "Irrigation" if "irrigation" in value else "Rain-fed"
            )
            if level == "High":
                frame = frame[frame["RCP"] == "RCP4.5"]
            selected = pd.concat(
                [
                    frame[frame["New Crop"].isin(main_codes)],
                    frame[frame["New Crop"].isin(other_codes)],
                ]
            )
            for _, row in selected.iterrows():
                suffix = row["Name"].split("_")[-1]
                variable = "cwd" if suffix == "wde" else "evt" if suffix == "eta" else suffix
                filename = f'{row["New Crop"]} {variable} {row["New Water Supply"]} {row["Input Level"]}.tif'
                final[filename] = {
                    "cache_filename": filename,
                    "model_crop_code": row["New Crop"],
                    "gaez_crop": row["Crop"],
                    "model_variable": variable,
                    "gaez_name": row["Name"],
                    "input_level": row["Input Level"],
                    "water_supply": row["New Water Supply"],
                    "available_water_content": row["Water Supply"],
                    "climate_model": row["Climate Model"],
                    "pathway": row["RCP"],
                    "time_period": row["Time Period"],
                    "source_unit": row["Data  Units"],
                    "download_url": row["Download URL"].strip(),
                    "file_identifier": str(row["File Identifier"]).lstrip("0") or "0",
                    "selection_table": table,
                    "selection_table_csv_line": int(row["_physical_line"]),
                }
    rows = [final[name] for name in sorted(final)]
    write_rows(GAEZ / "GAEZ_FJI_RASTER_CACHE_MANIFEST.csv", list(rows[0]), rows)


def reconstruct_ssp2() -> None:
    workbook_path = SSP / "iamc_db_POP_Countries.xlsx"
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["data"]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    selected_row = None
    excel_row = None
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(header, values))
        if (
            row.get("Model") == "IIASA-WiC POP"
            and row.get("Scenario") == "SSP2"
            and row.get("Region") == "FJI"
            and row.get("Variable") == "Population"
        ):
            selected_row, excel_row = row, number
            break
    workbook.close()
    if selected_row is None or excel_row is None:
        raise RuntimeError("SSP2 Fiji population row not found")
    years = list(range(2010, 2101, 5))
    extracted = {
        "workbook": workbook_path.name,
        "sheet": "data",
        "excel_row": excel_row,
        **{key: selected_row[key] for key in ("Model", "Scenario", "Region", "Variable", "Unit")},
        **{str(year): selected_row[year] for year in years},
    }
    write_rows(SSP / "SSP2_FJI_SELECTED_ROW.csv", list(extracted), [extracted])

    observed = {year: float(selected_row[year]) for year in years}
    annual: list[dict[str, object]] = []
    for year in range(2021, 2051):
        lower = year - year % 5
        upper = lower if year % 5 == 0 else lower + 5
        value = observed[lower] if lower == upper else observed[lower] + (
            observed[upper] - observed[lower]
        ) * (year - lower) / (upper - lower)
        annual.append({"year": year, "population_million": value})
    base = float(annual[0]["population_million"])
    for row in annual:
        row["population_index_2021"] = float(row["population_million"]) / base
    write_rows(SSP / "SSP2_FJI_ANNUAL_INDEX_2021_2050.csv", list(annual[0]), annual)


def main() -> int:
    main_codes, other_codes = reconstruct_faostat()
    reconstruct_gaez(main_codes, other_codes)
    reconstruct_ssp2()
    print(f"main crop codes: {main_codes}")
    print(f"additional crop codes: {other_codes}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
